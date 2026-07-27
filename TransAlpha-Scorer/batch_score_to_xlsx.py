import sys, os, socket
socket.setdefaulttimeout(8)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from transalpha.scoring.composite_scorer import CompositeScorer
from transalpha.scoring.position_sizer import PositionSizer

def code_to_suffix(code):
    c = code.strip().replace(".SH","").replace(".SZ","").replace(".BJ","")
    if c.startswith("6") or c.startswith("9"):
        return c + ".SH"
    return c + ".SZ"

with open("股票代码7.27.txt", encoding="utf-8") as f:
    raw_codes = [line.strip() for line in f if line.strip()]

codes = [code_to_suffix(c) for c in raw_codes]
print(f"共 {len(codes)} 只股票: {codes}")

scorer = CompositeScorer()
sizer = PositionSizer(total_capital=1000000, max_positions=5)

results = []
for code in codes:
    print(f"正在评分: {code} ...")
    try:
        r = scorer.calculate_composite_score(code)
        results.append(r)
        print(f"  {r.get('stock_name','')}: {'黑名单' if r['is_blacklisted'] else str(r.get('overall_score','N/A'))}")
    except Exception as e:
        print(f"  {code} 失败: {e}")
        results.append({"stock_code": code, "stock_name": "", "is_blacklisted": True, "error": str(e)})

suggestions = sizer.suggest_portfolio(results)
price_map = {}

wb = Workbook()
ws = wb.active
ws.title = "评分结果"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = [
    "股票代码", "股票名称", "所属行业", "综合得分", "评级",
    "PE分位打分", "PB分位打分", "ROE打分", "现金流打分",
    "增长稳定性打分", "盈利质量打分", "资产负债率打分", "价值基本面总分",
    "5日涨跌幅打分", "20日涨跌幅打分", "60日动量打分", "资金流入周期打分", "趋势动量总分",
    "宏观维度打分", "资金维度打分", "事件消息打分",
    "是否黑名单", "黑名单原因/风控预警",
    "建议仓位比例(%)", "建议持仓市值(元)", "建议持仓股数", "操作建议"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

data_font = Font(name="Arial", size=10)
data_align = Alignment(horizontal="center", vertical="center")
warn_font = Font(name="Arial", size=10, color="CC0000")

for i, r in enumerate(results, 2):
    code = r.get("stock_code", "")
    name = r.get("stock_name", "")
    industry = r.get("industry", "")
    total = r.get("overall_score")
    rating = r.get("rating", "")
    blacklisted = r.get("is_blacklisted", False)
    bl_reasons = "; ".join(r.get("blacklist_reasons", []))
    warnings = "; ".join(r.get("warnings", []))
    warn_text = bl_reasons or warnings or ""

    vf = r.get("value_fundamental", {})
    tm = r.get("trend_momentum", {})
    dim = r.get("dimensions", {})

    pos = None
    for s in suggestions:
        if s.get("stock_code") == code:
            pos = s
            break
    if not pos and not blacklisted and total is not None:
        price = price_map.get(code)
        pos = sizer.suggest_single_position(r, price)

    row_data = [
        code, name, industry,
        total if total is not None else "",
        rating,
        vf.get("pe_score", ""), vf.get("pb_score", ""), vf.get("roe_score", ""),
        vf.get("cash_flow_score", ""), vf.get("growth_stability_score", ""),
        vf.get("earnings_quality_score", ""), vf.get("debt_ratio_score", ""),
        vf.get("value_fundamental_score", ""),
        tm.get("five_day_return_score", ""), tm.get("twenty_day_return_score", ""),
        tm.get("sixty_day_momentum_score", ""), tm.get("fund_inflow_days_score", ""), tm.get("trend_momentum_score", ""),
        dim.get("macro_score", ""), dim.get("fund_flow_score", ""), dim.get("event_score", ""),
        "是" if blacklisted else "否",
        warn_text,
        pos.get("suggested_ratio", "") if pos else "",
        pos.get("suggested_value", "") if pos else "",
        pos.get("suggested_shares", "") if pos else "",
        pos.get("action", "") if pos else "",
    ]

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = warn_font if (blacklisted or warnings) and col == 22 else data_font
        cell.alignment = data_align
        cell.border = thin_border

col_widths = [14, 14, 16, 10, 10, 12, 12, 10, 12, 14, 12, 14, 14, 12, 14, 12, 14, 12, 12, 12, 12, 12, 30, 14, 14, 12, 12]
for col, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

ws.auto_filter.ref = ws.dimensions

output_path = "股票评分结果_7.27.xlsx"
wb.save(output_path)
print(f"\n结果已保存至: {os.path.abspath(output_path)}")
